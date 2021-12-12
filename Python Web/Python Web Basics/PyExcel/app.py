import openpyxl
#  for intellisense
from openpyxl.cell.cell import Cell

# wb = openpyxl.Workbook() new workbook

wb = openpyxl.load_workbook("transactions.xlsx")

print(wb.sheetnames)

sheet = wb["Sheet1"]

# wb.create_sheet("NewSheet", 0)
# wb.remove(sheet)

cell: Cell = sheet["a1"]

print(cell.row)

# sheet.cell(1, 1)
